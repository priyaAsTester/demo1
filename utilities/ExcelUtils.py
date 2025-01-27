import openpyxl

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def getColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    data=sheet.cell(row=rownum,column=columnno).value
    workbook.save(file)
    